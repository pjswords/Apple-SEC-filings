# DATA 620 Assignment 12.1
# Written by Pete Swords
# Last updated April 24, 2016

# This program retrieves and parses Apple's annual SEC filings from 2007 
# through 2015. After retrieving a report from Apple's website, the program 
# searches for the section titled "Item 1A. Risk Factors" and builds a 
# dictionary from the words in that section. The program then places 
# the contents of the dictionary into an object that corresponds to 
# a Microsoft Excel worksheet. While building the worksheets, the program also
# provides console output showing which words have the highest frequency.
# Once a worksheet is built for each report, the program saves them in a single
# XLSX file that can be used with Tableau.
#
# Note: This script requires access to the BeautifulSoup and OpenPyXL libraries.

import re, string, sys
from bs4 import BeautifulSoup
from openpyxl import Workbook
from urllib.error import URLError
from urllib.request import urlopen

# List of web pages for us to visit
webPageList = ("http://investor.apple.com/secfiling.cfm?filingID=1047469-07-9340&CIK=320193",
               "http://investor.apple.com/secfiling.cfm?filingID=1193125-08-224958&CIK=320193",
               "http://investor.apple.com/secfiling.cfm?filingID=1193125-09-214859&CIK=320193",
               "http://investor.apple.com/secfiling.cfm?filingID=1193125-10-238044&CIK=320193",
               "http://investor.apple.com/secfiling.cfm?filingID=1193125-11-282113&CIK=320193",
               "http://investor.apple.com/secfiling.cfm?filingID=1193125-12-444068&CIK=320193",
               "http://investor.apple.com/secfiling.cfm?filingID=1193125-13-416534&CIK=320193",
               "http://investor.apple.com/secfiling.cfm?filingID=1193125-14-383437&CIK=320193",
               "http://investor.apple.com/secfiling.cfm?filingID=1193125-15-356351&CIK=320193")

# Count will be used to keep track of which page we're on,
# and wb will be the OpenPyXL object that we'll save as an
# .xlsx file when we're done
count = 1
# A Workbook object to save our dictionaries to
wb = Workbook()
# A dictionary for our extracted words
d = dict()
# A (mostly) generic list of stopwords
stopWords = ("a","able","about","across","after","all","almost",
             "also","am","among","an","and","any","are","as",
             "at","be","because","been","but","by","can","cannot",
             "company","companys","could","dear","did","do","does","either",
             "else","ever","every","for","from","get","got","had","has",
             "have","he","her","hers","him","his","how","however",
             "i","if","in","into", "is","it","its","just","least",
             "let","like","likely","may","me","might","most",
             "must","my","neither","new","no","nor","not","of","off",
             "often","on","only","or","other","our","own","rather",
             "said","say","says","she","should","since","so","some",
             "such","than","that","the","their","them","then","there",
             "these","they","this","tis","to","too","twas","us",
             "wants","was","we","were","what","when","where","which",
             "while","who","whom","why","will","with","would","yet",
             "you","your")

# waitForUser(): Holds the console window open until user is ready to quit
def waitForUser():
    try:
        input("\nPress the Enter key to exit.")
        sys.exit()
    except KeyboardInterrupt:
        sys.exit()

# Iterate through the list of web pages
for page in webPageList:
    # A repository for the web page text
    pageText = ""
    
    # Try to open a new text file in write mode
    try:
        file = open("httpfile" + str(count) + ".txt", "wb")
    except IOError:
        print("Error: Could not open new text file for recording web page text. " +
              "Aborting program.")
        waitForUser()
        
    # Try to open a page and assign it to a handler
    try:
        fhand = urlopen(webPageList[count - 1]).read()
    except URLError:
        print("Error: Could not connect to website or unknown URL type specified. " +
              "Aborting program.")
        waitForUser()
        
    # Create a BeautifulSoup object from the handler
    soup = BeautifulSoup(fhand, 'html.parser')
    # Remove all script and style elements
    for script in soup(["script", "style"]):
        script.extract()
    # Assign what's left to a string
    pageText = soup.body.get_text()
    
    # Try to write the string to a file with UTF-8 encoding
    try:
        file.write(bytes(pageText, "utf-8"))
        file.close()
    except IOError:
        print("Error: Could not write to text file. Aborting program.")
        waitForUser()
    
    # Reset the repository
    pageText = ""
    # Try to open the text file we just created
    try:
        fhand = open("httpfile" + str(count) + ".txt", "r", -1 , "utf-8")
    except IOError:
        print("Error: Could not open file httpfile" + str(count) + ".txt" +
              ". Aborting program.")
        waitForUser()
        
    # This flag will be used in a bit to help us determine if we're in the right
    # location to start filling our dictionary
    flag = 0
    # Iterate through the text file
    for line in fhand:
        # Check for blank lines, skip to the next line when found
        if not line.strip():
            continue
        else:
            # Get rid of whitespace on each side of the line
            line = line.strip()
            # If the line ends with "Risk Factors", reset the dictionary and 
            # set flag to 1 -- as long as the next line doesn't start with a number,
            # we're in the right place.
            if re.search(r"Risk[\s\u00a0]*Factors$", line) and flag == 0:
                d.clear()
                print("Count " + str(count) + " of " + str(len(webPageList)) +
                      " - Item 1A found.")
                flag = 1
                continue
            # If the flag is set to 1 and the very next line begins with
            # a number, it means we're at the document TOC -- a false positive.
            # Set the flag to zero, go to the next line and keep looking. If flag > 1,
            # it means we're already at the target section.
            if re.search(r"^[0-9]", line) and flag == 1:
                print("False positive, continuing to search...")
                flag = 0
                continue
            # If we find "Item 1B." while the flag is set to anything greater than 1,
            # we can stop looping through lines and make our worksheet.
            if re.search(r"^Item[\s\u00a0]*1B\.", line) and flag > 1:
                print("Count " + str(count) + " of " + str(len(webPageList)) +
                      " - Item 1B found.")
                flag = 0
                break
            # If we've made it this far into the "else:" statement, we're filling
            # the dictionary. We'll keep doing this until we hit one of the
            # stop/reset conditions from above.
            if flag > 0:
                # The following print statement can be un-commented for debugging
                # print(line)
                flag = flag + 1
                # Get rid of string.punctuation, stopwords, and annoying unicode quotes
                line = "".join(c for c in line if c not in string.punctuation +
                               "\u201c\u201d\u0093\u0094\x92")
                line = line.lower()
                words = line.split()
                for word in words:
                    # Get rid of any words that are numbers or have numbers in them.
                    # Also, get rid of any words on our stopword list and any other
                    # words with less than 3 characters.
                    if re.search(r"[0-9]", word) or len(word) < 3 or \
                            word in(stopWords):
                        continue
                    if word not in d:
                        d[word] = 1
                    else:
                        d[word] = d[word] + 1

    # Use the active worksheet if we're on the first iteration,
    # otherwise create a new sheet.
    if count == 1:
        ws = wb.active
    else:
        ws = wb.create_sheet()
    # Give the sheet a meaningful title -- the year of the report
    ws.title = str(count + 2006)
    # Set an iterator to help us fill our worksheet
    i = 1
    # Add field labels to the worksheet
    ws["A" + str(i)] = "Year"
    ws["B" + str(i)] = "Word"
    ws["C" + str(i)] = "Count"
    # Put the dictionary into the workbook using the values as a key
    # to sort the dictionary from highest word count to lowest
    # While we're at it, we'll output the top 10 words to get an idea
    # of where we might focus our analytical efforts later.
    print("*** Top 10 words from this report ***")
    for x,y in sorted(d.items(), key = lambda x: x[1], reverse = True):
        i = i + 1
        ws["A" + str(i)] = count + 2006
        ws["B" + str(i)] = x
        ws["C" + str(i)] = y
        # Try to print top 10 words,
        # skip if bad unicode slipped past earlier checks
        try:
            if i < 12:
                print("\t" + x + " (" + str(y) + ")")
        except UnicodeEncodeError:
            continue
            
    # Let the user know when the worksheet is done
    print("Page " + str(count) + " of " + str(len(webPageList)) +
          " processed. Moving to next page...\n")
    
    # If we wanted to get rid of our text files, we could un-comment
    # the following lines and add "import os" to our script
    # try:
    #     os.remove("httpfile" + str(count) + ".txt")
    # except OSError:
    #     print("Could not remove httpfile" + str(count) + ".txt.")

    # Increment our count for the next iteration
    count = count + 1
    # Clear out the dictionary for the next iteration
    d.clear()

# When we've finished iterating through each web page and building
# its respective worksheet, save all worksheets to our new workbook
# and say goodbye.
try:
    wb.save("apple.xlsx")
except PermissionError:
    print("Error: Could not save the worksheet. Check to see if the target file" +
          "already exists and is open or flagged as read-only. Aborting program.")
    waitForUser()
print("Processing complete.")
waitForUser()

# End of script
